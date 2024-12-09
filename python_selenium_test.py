'''
撰寫者：Mika
撰寫日期：2023/05
使用工具：selenium
使用語言：python
主要目的：學習了python基礎入門課程後，嘗試能不能寫出自動化測試
備註：該駐點並未實現自動化測試項目，在沒有任何資源情況下自行嘗試。
     由於合約快要到期，因此，程式碼沒有完成。
     下方有GPT優化的版本，可以比較理解。
'''

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import os
import openpyxl

# 檔案路徑與檔名
path = r"C:\Users\ABCD"
file_name = r"xyz.xlsx"

os.chdir(path)
wb = openpyxl.load_workbook(file_name)
sheet = wb.worksheets[0]

# 開啟Chrom後，放大視窗
driver = webdriver.Chrome()
driver.maximize_window()

time.sleep(5)

# 輸入網址
driver.get("https://google.com")

# 帳號密碼登入
uername = driver.find_element(By.NAME, 'name_abc').send_keys("")
password = driver.find_element(By.NAME, 'password_abc').send_keys("")

# 執行登入按鈕
login = driver.find_element(By.TAG_NAME, 'button')
login.click()

time.sleep(3)

# 使用迴圈找尋管理頁面
title_name = driver.find_elements(By.XPATH, "//a[@class='dropdown-toggle ng-binding ng-scope']")
for Title_Name in range(len(title_name)) :
    if title_name[Title_Name].get_attribute("innerText").strip() == "管理頁面" :
        title_name_Action = ActionChains(driver)
        title_name_Action.move_to_element(title_name[Title_Name]).perform()

        # 找到總覽頁面並點擊
        sec_name = driver.find_elements(By.XPATH, "//a[@class='ng-binding ng-scope']")
        for Sec_Name in range(len(sec_name)) :
            if sec_name[Sec_Name].get_attribute("innerText").strip() == "總覽" :
                sec_name_Action = ActionChains(driver)
                sec_name_Action.move_to_element(sec_name[Sec_Name]).click(sec_name[Sec_Name]).perform()
                break

time.sleep(3)

# 找尋新增功能並點擊
add_button_name = driver.find_elements(By.CSS_SELECTOR, '[button-type="AdD"]')
for add_Button_Name in range(len(add_button_name)) :
    if add_button_name[add_Button_Name].get_attribute("innerText").strip() == '新增' :
        add_button_Action = ActionChains(driver)
        add_button_Action.move_to_element(add_button_name[add_Button_Name]).click(add_button_name[add_Button_Name]).perform()
        break

time.sleep(3)

# 當時想法是想從頁面尋找輸入欄位，並從Excel抓取資料。
row = 1
column = 2
one_class = driver.find_elements(By.CLASS_NAME, "col-lg-8")
for i in range(len(one_class)) :
    second_type = one_class[i].find_elements(By.CSS_SELECTOR, "input")
    second_tag = one_class[i].find_elements(By.TAG_NAME, "a")
    for n in range(len(second_type)) :                  

time.sleep(3)
driver.quit()


'''
======================
把另一個檔案整合在一起
======================
'''
import datetime
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import os
import re
from typing import Union

# 讀取檔案
path = r"C:\Users\ABCD"
file_name = r"xyz.xlsx"

os.chdir(path)
wb = openpyxl.load_workbook(file_name)
sheet = wb.worksheets[0]

def caseNum() :
    id_col = 1
    en_col = 1
    # 找尋編號以名稱(英文)
    for col in range(sheet.min_column, sheet.max_column + 1) :
        if sheet.cell(1, col).value == '編號'.strip() :
            id_col = col
        elif sheet.cell(1, col).value == '別名'.strip() :
            en_col = col
            break
    for row in range(2, sheet.max_row) :
        # 取得名稱(英文)
        en_value = sheet.cell(row, en_col).value
        if en_value is not None :
            # 取得名稱(英文)首/尾字
            en_simplify = en_value[0] + en_value[4] + en_value[-1]
          
            # 賦予產品編號
            productNum = "XX" + str.replace(str(datetime.datetime.today().date()), "-", "") + en_simplify
            sheet.cell(row, id_col).value = productNum
    wb.save(file_name)
caseNum()
print("Done!")




'''
===========================
以下是GPT優化的程式碼
===========================
'''
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time

# 設定檔案路徑與檔名
path = r"C:\Users\ABCD"
file_name = r"xyz.xlsx"

# 載入 Excel 檔案
wb = openpyxl.load_workbook(f"{path}\\{file_name}")
sheet = wb.active  # 使用第一個工作表

# 啟動 Selenium WebDriver
driver = webdriver.Chrome()
driver.maximize_window()

try:
    # 打開網站
    driver.get("https://example.com")
    time.sleep(3)

    # 登入
    username = driver.find_element(By.NAME, 'name_abc')
    password = driver.find_element(By.NAME, 'password_abc')
    username.send_keys("your_username")
    password.send_keys("your_password")
    driver.find_element(By.TAG_NAME, 'button').click()
    time.sleep(3)

    # 導航到目標頁面
    title_name = driver.find_elements(By.XPATH, "//a[@class='dropdown-toggle ng-binding ng-scope']")
    for title in title_name:
        if title.get_attribute("innerText").strip() == "管理頁面":
            ActionChains(driver).move_to_element(title).perform()
            break

    sec_name = driver.find_elements(By.XPATH, "//a[@class='ng-binding ng-scope']")
    for sec in sec_name:
        if sec.get_attribute("innerText").strip() == "總覽":
            ActionChains(driver).move_to_element(sec).click(sec).perform()
            break

    time.sleep(3)

    # 找到輸入欄位，依序填入 Excel 資料
    input_fields = driver.find_elements(By.CSS_SELECTOR, "input")
    for i, field in enumerate(input_fields):
        if i < sheet.max_row:  # 確保不超出 Excel 資料範圍
            value = sheet.cell(row=i+1, column=1).value  # 取 Excel 第一列資料
            field.send_keys(value if value else "")  # 填入資料，若為 None 則填空字串

    time.sleep(3)

finally:
    # 關閉瀏覽器
    driver.quit()




'''
============
另一段程式碼優化
============
'''
import datetime
import openpyxl
import os

# 讀取檔案
path = r"C:\Users\ABCD"
file_name = r"xyz.xlsx"

os.chdir(path)
wb = openpyxl.load_workbook(file_name)
sheet = wb.worksheets[0]

def caseNum():
    id_col = None
    en_col = None
    
    # 找尋欄位索引
    for col in range(sheet.min_column, sheet.max_column + 1):
        col_name = sheet.cell(1, col).value.strip().lower()
        if col_name == '編號'.strip().lower():
            id_col = col
        elif col_name == '別名'.strip().lower():
            en_col = col
            break

    if id_col is None or en_col is None:
        print("欄位名稱未找到！")
        return
    
    # 生成編號
    for row in range(2, sheet.max_row + 1):
        en_value = sheet.cell(row, en_col).value
        if en_value:
            # 確保名稱長度足夠
            if len(en_value) >= 5:
                # 取得名稱的首字、第四字、最後一字
                en_simplify = en_value[0] + en_value[3] + en_value[-1]
                # 生成產品編號
                product_num = "XX" + datetime.datetime.today().strftime("%Y%m%d") + en_simplify
                # 填入編號
                sheet.cell(row, id_col).value = product_num
            else:
                print(f"名稱 '{en_value}' 長度不足，無法生成編號！")

    wb.save(file_name)
    print("編號生成完成！")

caseNum()
